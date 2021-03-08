import os
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl import utils

from UiDataModel import TermCode, TerminologyEntry


def as_text(value):
    if value is None:
        return ""
    if "\n" in value:
        values = value.split()
        return max(values, key=len)
    return str(value)


def format_excel_sheet(sheet):
    for col in sheet.columns:
        for cell in col:
            cell.alignment = Alignment(wrapText=True, horizontal="left")
            length = max(len(as_text(cell.value)) for cell in col) + 5
            sheet.column_dimensions[utils.get_column_letter(col[0].column)].width = length


def get_termcode_row(terminology_code: TermCode):
    system = terminology_code.system
    code = terminology_code.code
    version = terminology_code.version
    display = terminology_code.display
    return [system, code, version, display]


def get_terminology_entry_row(terminology_entry: TerminologyEntry):
    result = ["UNDEFINED", "UNDEFINED", "UNDEFINED", "UNDEFINED"]
    if terminology_entry.termCode:
        result = get_termcode_row(terminology_entry.termCode)
    result.append(terminology_entry.display)
    return result


def to_excel(category_list):
    def add_terminology_entry(terminology_entry: TerminologyEntry):
        sheet.append(get_terminology_entry_row(terminology_entry))
        for child in terminology_entry.children:
            add_terminology_entry(child)
        for definition in terminology_entry.valueDefinitions:
            for concept in definition.selectableConcepts:
                sheet.append(get_termcode_row(concept))

    wb = Workbook()
    destination_filename = "FeasibilityUI-Display.xlsx"
    for category in category_list:
        sheet = wb.create_sheet(title=category.display.replace(" /", ""))
        for entry in category.children:
            add_terminology_entry(entry)
        format_excel_sheet(sheet)
    wb.save(filename=destination_filename)

